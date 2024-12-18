import argparse
import ast
import os
import time
import audit_config
from ai_engine import *
from project import ProjectAudit
from library.dataset_utils import load_dataset, Project
from planning import PlanningV2
from prompts import prompts
from sqlalchemy import create_engine
from dao import CacheManager, ProjectTaskMgr
import os
import pandas as pd
from openpyxl import Workbook,load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from codebaseQA.rag_processor import RAGProcessor
from res_processor.res_processor import ResProcessor

import dotenv
dotenv.load_dotenv()

def scan_project(project, db_engine):
    # 1. parsing projects  
    project_audit = ProjectAudit(project.id, project.path, db_engine)
    project_audit.parse(project.white_files, project.white_functions)
    
    #1.5 build rag
    rag_processor=RAGProcessor(project_audit.functions_to_check, "./src/codebaseQA/lancedb",project.id)
    # 2. planning & scanning
    project_taskmgr = ProjectTaskMgr(project.id, db_engine) 
    
    planning = PlanningV2(project_audit, project_taskmgr)
    # 
    engine = AiEngine(planning, project_taskmgr,rag_processor.db,"lancedb_"+project.id,project_audit)
    # 1. 扫描 
    engine.do_planning()
    engine.do_scan()

    return rag_processor.db,rag_processor.table_name,project_audit
    
    # 2. gpt4 对结果做rescan 
    # rescan_project_with_gpt4(project.id, db_engine)

def check_function_vul(engine,lancedb,lance_table_name,project_audit):
    project_taskmgr = ProjectTaskMgr(project.id, engine)
    engine = AiEngine(None, project_taskmgr,lancedb,lance_table_name,project_audit)
    engine.check_function_vul()
    # print(result)

def generate_excel(output_path, project_id):
    project_taskmgr = ProjectTaskMgr(project_id, engine)
    entities = project_taskmgr.query_task_by_project_id(project.id)
    
    # 创建一个空的DataFrame来存储所有实体的数据
    data = []
    for entity in entities:
        if "yes" in str(entity.result_gpt4).lower() and len(entity.business_flow_code)<=600:
            data.append({
                '漏洞结果': entity.result,
                'ID': entity.id,
                '项目名称': entity.project_id,
                '合同编号': entity.contract_code,
                'UUID': entity.key,
                '函数名称': entity.name,
                '函数代码': entity.content,
                '开始行': entity.start_line,
                '结束行': entity.end_line,
                '相对路径': entity.relative_file_path,
                '绝对路径': entity.absolute_file_path,
                '业务流程代码': entity.business_flow_code,
                '业务流程行': entity.business_flow_lines,
                '业务流程上下文': entity.business_flow_context,
                '确认结果': entity.result_gpt4,
                '确认细节': entity.category
            })
    
    # 将数据转换为DataFrame
    if not data:  # 检查是否有数据
        print("No data to process")
        return
        
    df = pd.DataFrame(data)
    
    try:
        # 对df进行漏洞归集处理
        res_processor = ResProcessor(df)
        processed_df = res_processor.process()
        
        # 确保所有必需的列都存在
        required_columns = df.columns
        for col in required_columns:
            if col not in processed_df.columns:
                processed_df[col] = ''
                
        # 重新排列列顺序以匹配原始DataFrame
        processed_df = processed_df[df.columns]
    except Exception as e:
        print(f"Error processing data: {e}")
        processed_df = df  # 如果处理失败，使用原始DataFrame
    
    # 确保输出目录存在
    output_dir = os.path.dirname(output_path)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # 检查文件是否存在，如果不存在则创建新文件
    if not os.path.exists(output_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "项目数据"
    else:
        wb = load_workbook(output_path)
        if "项目数据" in wb.sheetnames:
            ws = wb["项目数据"]
        else:
            ws = wb.create_sheet("项目数据")
    
    # 如果工作表是空的，添加表头
    if ws.max_row == 1:
        for col, header in enumerate(processed_df.columns, start=1):
            ws.cell(row=1, column=col, value=header)
    
    # 将DataFrame数据写入工作表
    for row in dataframe_to_rows(processed_df, index=False, header=False):
        ws.append(row)
    
    # 保存Excel文件
    wb.save(output_path)
    
    print(f"Excel文件已保存到: {output_path}")
if __name__ == '__main__':

    switch_production_or_test = 'test' # prod / test

    if switch_production_or_test == 'test':
        start_time=time.time()
        db_url_from = os.environ.get("DATABASE_URL")
        engine = create_engine(db_url_from)
        
        dataset_base = "./src/dataset/agent-v1-c4"
        projects = load_dataset(dataset_base)

        project_id = 'chillz111'
        project_path = ''
        project = Project(project_id, projects[project_id])
        
        cmd = 'detect_vul'
        if cmd == 'detect_vul':
            lancedb,lance_table_name,project_audit=scan_project(project, engine) # scan
            check_function_vul(engine,lancedb,lance_table_name,project_audit) # confirm
        # elif cmd == 'check_vul_if_positive':
        #     check_function_vul(engine) # confirm

        end_time=time.time()
        print("Total time:",end_time-start_time)
        generate_excel("./output.xlsx",project_id)
        
        
    if switch_production_or_test == 'prod':
        # Set up command line argument parsing
        parser = argparse.ArgumentParser(description='Process input parameters for vulnerability scanning.')
        parser.add_argument('-fpath', type=str, required=True, help='Combined base path for the dataset and folder')
        parser.add_argument('-id', type=str, required=True, help='Project ID')
        # parser.add_argument('-cmd', type=str, choices=['detect', 'confirm','all'], required=True, help='Command to execute')
        parser.add_argument('-o', type=str, required=True, help='Output file path')
        # usage:
        # python main.py 
        # --fpath ../../dataset/agent-v1-c4/Archive 
        # --id Archive_aaa 
        # --cmd detect

        # Parse arguments
        args = parser.parse_args()
        print("fpath:",args.fpath)
        print("id:",args.id)
        print("cmd:",args.cmd)
        print("o:",args.o)
        # Split dataset_folder into dataset and folder
        dataset_base, folder_name = os.path.split(args.fpath)
        print("dataset_base:",dataset_base)
        print("folder_name:",folder_name)
        # Start time
        start_time = time.time()

        # Database setup
        db_url_from = os.environ.get("DATABASE_URL")
        engine = create_engine(db_url_from)

        # Load projects
        projects = load_dataset(dataset_base, args.id, folder_name)
        project = Project(args.id, projects[args.id])

        # Execute command
        # if args.cmd == 'detect':
        #     scan_project(project, engine)  # scan            
        # elif args.cmd == 'confirm':
        #     check_function_vul(engine)  # confirm
        # elif args.cmd == 'all':
        lancedb=scan_project(project, engine)  # scan
        check_function_vul(engine,lancedb)  # confirm

        end_time = time.time()
        print("Total time:", end_time -start_time)
        generate_excel(args.o,args.id)




